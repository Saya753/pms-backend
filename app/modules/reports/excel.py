from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# COLORS / STYLES
# =========================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SUBHEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

SUMMARY_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

DELAY_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

BOLD_FONT = Font(
    bold=True,
)

NORMAL_FONT = Font(
    color="000000",
)

THIN_SIDE = Side(
    style="thin",
    color="B7B7B7",
)

THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


# =========================================================
# GENERAL HELPERS
# =========================================================

def apply_header_style(ws, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER


def apply_table_borders(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
):
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )


def auto_width(ws, minimum: int = 10, maximum: int = 35):
    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is None:
                continue

            value = str(cell.value)

            max_length = max(
                max_length,
                len(value),
            )

        ws.column_dimensions[
            column_letter
        ].width = max(
            minimum,
            min(max_length + 2, maximum),
        )


def safe_sheet_name(
    name: str,
    used_names: set[str],
) -> str:

    invalid_chars = [
        "\\",
        "/",
        "*",
        "?",
        ":",
        "[",
        "]",
    ]

    clean_name = str(name)

    for char in invalid_chars:
        clean_name = clean_name.replace(
            char,
            "_",
        )

    clean_name = clean_name.strip()

    if not clean_name:
        clean_name = "پروژه"

    # Excel sheet name max length = 31
    clean_name = clean_name[:31]

    base_name = clean_name
    counter = 1

    while clean_name in used_names:

        suffix = f"_{counter}"

        clean_name = (
            base_name[:31 - len(suffix)]
            + suffix
        )

        counter += 1

    used_names.add(clean_name)

    return clean_name


def format_date(value):
    if value is None:
        return "-"

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    return str(value)


# =========================================================
# PROJECT CALCULATIONS
# =========================================================

def calculate_project_progress(
    project_id: int,
    tasks,
) -> float:

    main_tasks = [
        task
        for task in tasks
        if task.project_id == project_id
        and task.parent_id is None
    ]

    if not main_tasks:
        return 0.0

    progress = (
        sum(task.progress for task in main_tasks)
        / len(main_tasks)
    )

    return round(progress, 2)


def calculate_project_delay(
    project,
) -> int:

    if project.end_date is None:
        return 0

    if project.status in {
        "DONE",
        "COMPLETED",
        "CANCELLED",
    }:
        return 0

    today = date.today()

    if project.end_date >= today:
        return 0

    return (
        today - project.end_date
    ).days


# =========================================================
# SUMMARY SHEET
# =========================================================

def create_summary_sheet(
    wb,
    report,
):

    ws = wb.active
    ws.title = "Summary"

    ws.sheet_view.rightToLeft = True

    ws["A1"] = "گزارش کلی مدیریت پروژه"
    ws["A1"].font = Font(
        bold=True,
        size=18,
    )

    ws.merge_cells(
        "A1:B1"
    )

    ws["A2"] = "تاریخ تهیه گزارش"
    ws["B2"] = date.today().isoformat()

    ws["A4"] = "شاخص"
    ws["B4"] = "مقدار"

    apply_header_style(
        ws,
        row=4,
        start_col=1,
        end_col=2,
    )

    summary = report.summary

    rows = [
        (
            "تعداد کل پروژه‌ها",
            summary.total_projects,
        ),
        (
            "تعداد کل فعالیت‌ها",
            summary.total_tasks,
        ),
        (
            "فعالیت‌های تکمیل‌شده",
            summary.completed_tasks,
        ),
        (
            "نرخ تکمیل فعالیت‌ها",
            summary.completed_tasks
            / summary.total_tasks
            if summary.total_tasks
            else 0,
        ),
        (
            "میانگین پیشرفت پروژه‌ها",
            summary.average_project_progress
            / 100,
        ),
    ]

    for index, (label, value) in enumerate(
        rows,
        start=5,
    ):

        ws.cell(
            row=index,
            column=1,
            value=label,
        )

        ws.cell(
            row=index,
            column=2,
            value=value,
        )

        ws.cell(
            row=index,
            column=1,
        ).font = BOLD_FONT

        ws.cell(
            row=index,
            column=1,
        ).fill = SUMMARY_FILL

        ws.cell(
            row=index,
            column=2,
        ).fill = SUMMARY_FILL

        ws.cell(
            row=index,
            column=1,
        ).border = THIN_BORDER

        ws.cell(
            row=index,
            column=2,
        ).border = THIN_BORDER

    # -------------------------------------------------
    # Extra management indicators
    # -------------------------------------------------

    active_projects = sum(
        1
        for project in report.projects
        if project.status not in {
            "DONE",
            "COMPLETED",
            "CANCELLED",
        }
    )

    completed_projects = sum(
        1
        for project in report.projects
        if project.status in {
            "DONE",
            "COMPLETED",
        }
    )

    delayed_projects = sum(
        1
        for project in report.projects
        if project.is_delayed
    )

    extra_start = 10

    extra_rows = [
        (
            "پروژه‌های فعال",
            active_projects,
        ),
        (
            "پروژه‌های تکمیل‌شده",
            completed_projects,
        ),
        (
            "پروژه‌های دارای تأخیر",
            delayed_projects,
        ),
        (
            "نرخ تکمیل فعالیت‌ها",
            report.performance.task_completion_rate
            / 100,
        ),
        (
            "میانگین پیشرفت پروژه‌ها",
            report.performance.average_project_progress
            / 100,
        ),
    ]

    for index, (label, value) in enumerate(
        extra_rows,
        start=extra_start,
    ):

        ws.cell(
            row=index,
            column=1,
            value=label,
        )

        ws.cell(
            row=index,
            column=2,
            value=value,
        )

        ws.cell(
            row=index,
            column=1,
        ).font = BOLD_FONT

        ws.cell(
            row=index,
            column=1,
        ).fill = SUBHEADER_FILL

        ws.cell(
            row=index,
            column=2,
        ).fill = SUBHEADER_FILL

        ws.cell(
            row=index,
            column=1,
        ).border = THIN_BORDER

        ws.cell(
            row=index,
            column=2,
        ).border = THIN_BORDER

    # Percentage formatting
    for row in [
        8,
        9,
        13,
        14,
    ]:

        ws.cell(
            row=row,
            column=2,
        ).number_format = "0.00%"

    ws.freeze_panes = "A5"

    auto_width(ws)


# =========================================================
# PROJECT PORTFOLIO SHEET
# =========================================================

def create_project_portfolio_sheet(
    wb,
    report,
    projects,
    tasks,
):

    ws = wb.create_sheet(
        "سبد پروژه‌ها"
    )

    ws.sheet_view.rightToLeft = True

    ws["A1"] = "سبد پروژه‌ها"

    ws["A1"].font = Font(
        bold=True,
        size=18,
    )

    headers = [
        "ردیف",
        "نام پروژه",
        "وضعیت",
        "اولویت",
        "شروع برنامه‌ای",
        "پایان برنامه‌ای",
        "شروع واقعی",
        "پایان واقعی",
        "تأخیر (روز)",
        "پیشرفت برنامه‌ای",
        "پیشرفت واقعی",
        "تعداد فعالیت",
        "فعالیت تکمیل‌شده",
    ]

    header_row = 3

    for col, header in enumerate(
        headers,
        start=1,
    ):

        ws.cell(
            row=header_row,
            column=col,
            value=header,
        )

    apply_header_style(
        ws,
        row=header_row,
        start_col=1,
        end_col=len(headers),
    )

    for index, project in enumerate(
        projects,
        start=1,
    ):

        project_tasks = [
            task
            for task in tasks
            if task.project_id == project.id
        ]

        completed_tasks = sum(
            1
            for task in project_tasks
            if task.status == "DONE"
        )

        progress = calculate_project_progress(
            project.id,
            tasks,
        )

        delay = calculate_project_delay(
            project,
        )

        row = header_row + index

        values = [
            index,
            project.name,
            project.status,
            getattr(project, "priority", None) or "-",
            format_date(
                getattr(
                    project,
                    "start_date",
                    None,
                )
            ),
            format_date(
                getattr(
                    project,
                    "end_date",
                    None,
                )
            ),
            "-",
            "-",
            delay,
            "-",
            progress / 100,
            len(project_tasks),
            completed_tasks,
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):

            ws.cell(
                row=row,
                column=col,
                value=value,
            )

        if delay > 0:

            for col in range(
                1,
                len(headers) + 1,
            ):

                ws.cell(
                    row=row,
                    column=col,
                ).fill = DELAY_FILL

    last_row = (
        header_row
        + len(projects)
    )

    if projects:
        apply_table_borders(
            ws,
            start_row=header_row,
            end_row=last_row,
            start_col=1,
            end_col=len(headers),
        )

    # Percentage columns
    for row in range(
        header_row + 1,
        last_row + 1,
    ):

        ws.cell(
            row=row,
            column=10,
        ).number_format = "0%"

        ws.cell(
            row=row,
            column=11,
        ).number_format = "0.00%"

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = (
        f"A3:M{max(last_row, 3)}"
    )

    auto_width(ws)


# =========================================================
# ACTIVITIES DATABASE
# =========================================================

def create_activity_database_sheet(
    wb,
    projects,
    tasks,
):

    ws = wb.create_sheet(
        "Database_فعالیت‌ها"
    )

    ws.sheet_view.rightToLeft = True

    ws["A1"] = "Database فعالیت‌ها"

    ws["A1"].font = Font(
        bold=True,
        size=18,
    )

    headers = [
        "ردیف",
        "پروژه",
        "فعالیت",
        "نوع فعالیت",
        "وضعیت",
        "اولویت",
        "شروع",
        "پایان",
        "پیشرفت",
        "شناسه مسئول",
    ]

    header_row = 3

    for col, header in enumerate(
        headers,
        start=1,
    ):

        ws.cell(
            row=header_row,
            column=col,
            value=header,
        )

    apply_header_style(
        ws,
        row=header_row,
        start_col=1,
        end_col=len(headers),
    )

    project_names = {
        project.id: project.name
        for project in projects
    }

    sorted_tasks = sorted(
        tasks,
        key=lambda task: (
            project_names.get(
                task.project_id,
                "",
            ),
            task.created_at,
        ),
    )

    for index, task in enumerate(
        sorted_tasks,
        start=1,
    ):

        row = header_row + index

        activity_type = (
            "فعالیت اصلی"
            if task.parent_id is None
            else "زیر فعالیت"
        )

        values = [
            index,
            project_names.get(
                task.project_id,
                f"Project {task.project_id}",
            ),
            task.title,
            activity_type,
            task.status,
            task.priority,
            format_date(
                task.start_date
            ),
            format_date(
                task.due_date
            ),
            task.progress / 100,
            task.assignee_id
            if task.assignee_id is not None
            else "-",
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):

            ws.cell(
                row=row,
                column=col,
                value=value,
            )

    last_row = (
        header_row
        + len(sorted_tasks)
    )

    if sorted_tasks:
        apply_table_borders(
            ws,
            start_row=header_row,
            end_row=last_row,
            start_col=1,
            end_col=len(headers),
        )

    for row in range(
        header_row + 1,
        last_row + 1,
    ):

        ws.cell(
            row=row,
            column=9,
        ).number_format = "0%"

    ws.freeze_panes = "A4"

    ws.auto_filter.ref = (
        f"A3:J{max(last_row, 3)}"
    )

    auto_width(ws)


# =========================================================
# INDIVIDUAL PROJECT SHEETS
# =========================================================

def create_project_sheet(
    wb,
    project,
    tasks,
    used_sheet_names,
):

    sheet_name = safe_sheet_name(
        project.name,
        used_sheet_names,
    )

    ws = wb.create_sheet(
        sheet_name
    )

    ws.sheet_view.rightToLeft = True

    project_tasks = [
        task
        for task in tasks
        if task.project_id == project.id
    ]

    main_tasks = [
        task
        for task in project_tasks
        if task.parent_id is None
    ]

    progress = (
        sum(task.progress for task in main_tasks)
        / len(main_tasks)
        if main_tasks
        else 0
    )

    completed_tasks = sum(
        1
        for task in project_tasks
        if task.status == "DONE"
    )

    delay = calculate_project_delay(
        project
    )

    # -------------------------------------------------
    # Project header
    # -------------------------------------------------

    ws["A1"] = project.name

    ws["A1"].font = Font(
        bold=True,
        size=18,
    )

    ws.merge_cells(
        "A1:D1"
    )

    summary_rows = [
        (
            "وضعیت پروژه",
            project.status,
        ),
        (
            "اولویت",
            getattr(
                project,
                "priority",
                None,
            ) or "-",
        ),
        (
            "شروع برنامه‌ای",
            format_date(
                getattr(
                    project,
                    "start_date",
                    None,
                )
            ),
        ),
        (
            "پایان برنامه‌ای",
            format_date(
                getattr(
                    project,
                    "end_date",
                    None,
                )
            ),
        ),
        (
            "شروع واقعی",
            "-",
        ),
        (
            "پایان واقعی",
            "-",
        ),
        (
            "تأخیر",
            f"{delay} روز",
        ),
        (
            "پیشرفت واقعی",
            progress / 100,
        ),
        (
            "تعداد فعالیت‌ها",
            len(project_tasks),
        ),
        (
            "فعالیت‌های تکمیل‌شده",
            completed_tasks,
        ),
    ]

    for index, (label, value) in enumerate(
        summary_rows,
        start=3,
    ):

        ws.cell(
            row=index,
            column=1,
            value=label,
        )

        ws.cell(
            row=index,
            column=2,
            value=value,
        )

        ws.cell(
            row=index,
            column=1,
        ).font = BOLD_FONT

        ws.cell(
            row=index,
            column=1,
        ).fill = SUBHEADER_FILL

        ws.cell(
            row=index,
            column=1,
        ).border = THIN_BORDER

        ws.cell(
            row=index,
            column=2,
        ).border = THIN_BORDER

    ws["B10"].number_format = "0.00%"

    if delay > 0:

        ws["B9"].fill = DELAY_FILL

    # -------------------------------------------------
    # Tasks table
    # -------------------------------------------------

    table_start = 15

    headers = [
        "ردیف",
        "فعالیت",
        "نوع",
        "وضعیت",
        "اولویت",
        "شروع",
        "پایان",
        "پیشرفت",
        "مسئول",
        "شناسه مسئول",
    ]

    for col, header in enumerate(
        headers,
        start=1,
    ):

        ws.cell(
            row=table_start,
            column=col,
            value=header,
        )

    apply_header_style(
        ws,
        row=table_start,
        start_col=1,
        end_col=len(headers),
    )

    for index, task in enumerate(
        project_tasks,
        start=1,
    ):

        row = table_start + index

        activity_type = (
            "فعالیت اصلی"
            if task.parent_id is None
            else "زیر فعالیت"
        )

        values = [
            index,
            task.title,
            activity_type,
            task.status,
            task.priority,
            format_date(
                task.start_date
            ),
            format_date(
                task.due_date
            ),
            task.progress / 100,
            task.assignee_id
            if task.assignee_id is not None
            else "-",
            task.assignee_id
            if task.assignee_id is not None
            else "-",
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):

            ws.cell(
                row=row,
                column=col,
                value=value,
            )

    last_row = (
        table_start
        + len(project_tasks)
    )

    if project_tasks:
        apply_table_borders(
            ws,
            start_row=table_start,
            end_row=last_row,
            start_col=1,
            end_col=len(headers),
        )

    for row in range(
        table_start + 1,
        last_row + 1,
    ):

        ws.cell(
            row=row,
            column=8,
        ).number_format = "0%"

    ws.freeze_panes = "A16"

    ws.auto_filter.ref = (
        f"A15:J{max(last_row, 15)}"
    )

    auto_width(ws)


# =========================================================
# MAIN EXCEL CREATOR
# =========================================================

def create_general_report_excel(
    report,
    projects,
    tasks,
):

    wb = Workbook()

    # -------------------------------------------------
    # Summary
    # -------------------------------------------------

    create_summary_sheet(
        wb=wb,
        report=report,
    )

    # -------------------------------------------------
    # Portfolio
    # -------------------------------------------------

    create_project_portfolio_sheet(
        wb=wb,
        report=report,
        projects=projects,
        tasks=tasks,
    )

    # -------------------------------------------------
    # Activity database
    # -------------------------------------------------

    create_activity_database_sheet(
        wb=wb,
        projects=projects,
        tasks=tasks,
    )

    # -------------------------------------------------
    # Project sheets
    # -------------------------------------------------

    used_sheet_names = {
        "Summary",
        "سبد پروژه‌ها",
        "Database_فعالیت‌ها",
    }

    # Find actual project objects.
    # report.projects contains report information,
    # so project information needed for detailed sheets
    # is reconstructed from report data.

    for project in projects:

        create_project_sheet(
            wb=wb,
            project=project,
            tasks=tasks,
            used_sheet_names=used_sheet_names,
    )

    # -------------------------------------------------
    # Workbook settings
    # -------------------------------------------------

    wb.active = 0

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output


def create_project_report_excel(
    report,
    project,
    tasks,
):
    wb = Workbook()

    # -----------------------------
    # Summary
    # -----------------------------

    create_summary_sheet(
        wb=wb,
        report=report,
    )

    # -----------------------------
    # Only this project's tasks
    # -----------------------------

    project_tasks = [
        task
        for task in tasks
        if task.project_id == project.id
    ]

    # -----------------------------
    # Project sheet
    # -----------------------------

    used_sheet_names = {
        "Summary",
    }

    create_project_sheet(
        wb=wb,
        project=project,
        tasks=project_tasks,
        used_sheet_names=used_sheet_names,
    )

    # -----------------------------
    # Activity database
    # -----------------------------

    create_activity_database_sheet(
        wb=wb,
        projects=[project],
        tasks=project_tasks,
    )

    wb.active = 0

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output